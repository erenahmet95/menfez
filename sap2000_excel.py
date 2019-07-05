from openpyxl import *
def excel(deger,yol):
    yol=yol

    tek_eksenli = load_workbook(yol)
    sayfa = tek_eksenli.active
    max_satır_sayısı = sayfa.max_row
    max_sutun_sayısı = sayfa.max_column
    min_satır_sayısı = sayfa.min_row
    min_sutun_sayısı = sayfa.min_column
    matris_satır=[]
    matris_sutun=[]
    baslangıc=deger




    for satır in sayfa.iter_rows(min_row=min_satır_sayısı,min_col=(baslangıc-1),max_row=max_satır_sayısı,max_col=(baslangıc-1)):


        for hücre in satır :


            matris_satır.append(hücre.value)


    #devam=input(" Listelerde Stringlerin Silinmesini İstiyorsanız 'a' ya İStemiyorsanız her hangi bir tuşa tıklayın  ")

    #if devam == "a":
        #for i in matris_satır:

            #if type(i) == str :
                #print(" Satırlar Arasında String Eleman olduğu İçin Silindi")
                #matris_satır.remove(i)

    print("------ Koordinat Excell Verileri Aktarıldı------")
    return list(matris_satır)

def  excel2():
    yol=input("Dosya : ")

    import os
    if os.path.exists(yol):

        genel_matris=[]
        tek_eksenli = load_workbook(yol)
        sayfa = tek_eksenli.active
        max_sutun_sayısı = sayfa.max_column
        for i in range(2,max_sutun_sayısı+2):

            genel_matris.append(excel(i,yol))


        return genel_matris
    else :
        print("Böyle bir Dosya Yok ")
